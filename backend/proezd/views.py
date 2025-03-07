import os
import pandas as pd
import re
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.db import connection
from django.http import JsonResponse, FileResponse
from .forms import PotokUploadForm, PropuskUploadForm
from datetime import datetime, timedelta
import logging
from difflib import SequenceMatcher
from collections import defaultdict
from functools import lru_cache
import json
import multiprocessing as mp
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from psycopg2.extras import execute_values
import zipfile

POTOK_TABLE = os.getenv("POSTGRES_TABLE_POTOK")
PROPUSK_TABLE = os.getenv("POSTGRES_TABLE_PROPUSK")
COMPANY_TABLE = os.getenv("POSTGRES_TABLE_COMPANY")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler()],
)
logger = logging.getLogger(__name__)


@lru_cache(maxsize=1000)
def similar(a, b):
    """Вычисляет схожесть двух строк с учетом форматов российских номеров"""

    def escape_pattern(s):
        # Экранируем все специальные символы regex, кроме точки (для ?)
        special_chars = "[](){}*+?^$|\\"
        return "".join("\\" + c if c in special_chars else c for c in s)

    # Заменяем ? на . и экранируем остальные спецсимволы
    a_pattern = escape_pattern(a.replace("?", "."))
    b_pattern = escape_pattern(b.replace("?", "."))

    # Если в номере есть ?, считаем его частично распознанным
    if "?" in a or "?" in b:
        if re.match(f"^{a_pattern}$", b) or re.match(f"^{b_pattern}$", a):
            return 0.9

    # Быстрая проверка на полное совпадение
    if a == b:
        return 1.0

    # Быстрая проверка на длину
    if abs(len(a) - len(b)) > 2:
        return 0.0

    # Оптимизированная проверка на схожие символы
    similar_digits = {"0": "О", "8": "В", "3": "З"}
    if len(a) == len(b):
        matches = sum(
            1
            for i in range(len(a))
            if a[i] == b[i]
            or a[i] == "?"
            or b[i] == "?"
            or (a[i] in similar_digits and b[i] == similar_digits[a[i]])
            or (b[i] in similar_digits and a[i] == similar_digits[b[i]])
        )
        return matches / len(a)

    # Для разной длины используем базовое сравнение
    return SequenceMatcher(None, a, b).ratio()


def prepare_reference_numbers(numbers):
    by_length = defaultdict(set)
    by_prefix = defaultdict(set)
    for num in numbers:
        length = len(num)
        by_length[length].add(num)
        if len(num) >= 2:
            prefix = num[:2]
            by_prefix[prefix].add(num)
    return by_length, by_prefix


def get_most_similar_number(plate, ref_data, threshold=0.6):
    by_length, by_prefix = ref_data
    plate_len = len(plate)

    # Быстрая проверка на точное совпадение
    if plate in by_length.get(plate_len, set()):
        return (plate, 1.0)

    # Проверяем только номера похожей длины
    candidates = set()
    for length in range(max(4, plate_len - 1), min(10, plate_len + 2)):
        candidates.update(by_length.get(length, set()))

    # Если есть префикс, фильтруем по нему
    if len(plate) >= 2:
        prefix = plate[:2]
        prefix_matches = by_prefix.get(prefix, set())
        if prefix_matches:
            candidates = candidates.intersection(prefix_matches)
            if not candidates and "?" in prefix:
                # Если префикс содержит ?, проверяем все номера подходящей длины
                candidates = set().union(
                    *(
                        by_length.get(l, set())
                        for l in range(max(4, plate_len - 1), min(10, plate_len + 2))
                    )
                )

    best_match = None
    best_similarity = threshold

    for ref in candidates:
        # Быстрая проверка на общие символы
        common_chars = sum(
            1 for a, b in zip(plate, ref) if a == b or a == "?" or b == "?"
        )
        if common_chars / max(len(plate), len(ref)) < threshold:
            continue

        similarity = similar(plate, ref)
        if similarity > best_similarity:
            best_similarity = similarity
            best_match = (ref, similarity)
            # Если нашли почти полное совпадение, прекращаем поиск
            if similarity > 0.95:
                break

    return best_match if best_match and best_match[1] >= 0.88 else None


def process_chunk(chunk_data):
    """Обработка части номеров в отдельном процессе"""
    try:
        chunk_plates, ref_data, threshold = chunk_data
        results = []

        # Создаем множество эталонных номеров для быстрой проверки
        all_ref_numbers = set()
        for numbers in ref_data[0].values():
            all_ref_numbers.update(numbers)

        # Предварительная фильтрация номеров с оптимизацией
        filtered_plates = []
        ru_letters = set("АВЕКМНОРСТУХ")

        for plate, dt, potok_id in chunk_plates:
            try:
                if not plate:
                    continue

                plate = str(plate).upper()
                # Быстрая проверка на недопустимые буквы
                if any(c.isalpha() and c not in ru_letters for c in plate):
                    continue

                # Быстрая проверка на точное совпадение
                if plate in all_ref_numbers:
                    continue

                filtered_plates.append((plate, dt, potok_id))
            except Exception as e:
                logger.error(f"Ошибка при обработке номера {plate}: {str(e)}")
                continue

        # Оптимизированный поиск похожих номеров
        for plate, dt, potok_id in filtered_plates:
            try:
                # Поиск похожих номеров с оптимизированным порогом
                match_result = get_most_similar_number(
                    plate, ref_data, 0.5 if "?" in plate else threshold
                )

                if match_result:
                    ref_num, similarity = match_result
                    if similarity >= 0.88:
                        results.append(
                            {
                                "id": potok_id,
                                "original": plate,
                                "suggested": ref_num,
                                "similarity": f"{similarity:.2%}",
                                "dt": dt.strftime("%Y-%m-%d %H:%M:%S"),
                            }
                        )
            except Exception as e:
                logger.error(f"Ошибка при поиске похожих номеров для {plate}: {str(e)}")
                continue

        return results
    except Exception as e:
        logger.error(f"Критическая ошибка в process_chunk: {str(e)}")
        return []


@staff_member_required
def analyze_numbers(request):
    try:
        logger.info(f"Получен {request.method} запрос на анализ номеров")
        logger.info(f"Content-Type: {request.headers.get('Content-Type')}")
        logger.info(f"X-CSRFToken: {request.headers.get('X-CSRFToken', 'Нет')}")

        if request.method == "GET":
            logger.info("Возвращаем страницу анализа номеров")
            return render(request, "admin/analyze_numbers.html")

        if request.GET.get("sort") == "true":
            results = request.session.get("analysis_results", [])
            if not results:
                logger.warning("Нет сохраненных результатов анализа")
                return JsonResponse(
                    {"error": "Необходимо выполнить анализ заново"}, status=400
                )

            sort_ascending = (
                request.GET.get("sort_ascending", "false").lower() == "true"
            )
            results.sort(
                key=lambda x: float(x["similarity"].rstrip("%")),
                reverse=not sort_ascending,
            )
            logger.info(
                f"Результаты отсортированы по {'возрастанию' if sort_ascending else 'убыванию'} процента схожести"
            )

            return JsonResponse({"results": results})

        logger.info("Начинаем анализ номеров...")

        # Уменьшаем размер пакета для более стабильной работы
        BATCH_SIZE = 500

        with connection.cursor() as cursor:
            try:
                logger.info("Получаем список эталонных номеров из базы...")
                cursor.execute(
                    f"""
                    SELECT gn
                    FROM {PROPUSK_TABLE}
                    WHERE gn IS NOT NULL
                    AND dateactual IS NULL
                """
                )
                reference_numbers = {row[0].upper() for row in cursor.fetchall()}
                logger.info(f"Получено {len(reference_numbers)} эталонных номеров")

                if not reference_numbers:
                    logger.warning("Не найдено эталонных номеров в базе")
                    return JsonResponse(
                        {"error": "Не найдено эталонных номеров в базе"}, status=400
                    )

                ref_data = prepare_reference_numbers(reference_numbers)
                logger.info(
                    f"Подготовлены индексы для поиска. Длины номеров: {len(ref_data[0])}, Префиксы: {len(ref_data[1])}"
                )

                cursor.execute(
                    f"""
                    SELECT COUNT(*)
                    FROM {POTOK_TABLE}
                    WHERE gosnmr IS NOT NULL
                    AND del IS NULL
                    AND gosnmr !~ '[A-Za-z]'
                """
                )
                total_records = cursor.fetchone()[0]
                logger.info(f"Всего записей для обработки: {total_records}")

                # Инициализируем список для хранения результатов
                all_results = []
                processed_records = 0

                # Ограничиваем количество процессов и добавляем тайм-аут
                num_processes = min(mp.cpu_count(), 2)
                logger.info(f"Используем {num_processes} процессов для обработки")

                while processed_records < total_records:
                    try:
                        cursor.execute(
                            f"""
                            SELECT gosnmr, dt, potok_id
                            FROM {POTOK_TABLE}
                            WHERE gosnmr IS NOT NULL
                            AND del IS NULL
                            AND gosnmr !~ '[A-Za-z]'
                            ORDER BY dt
                            LIMIT {BATCH_SIZE}
                            OFFSET {processed_records}
                        """
                        )

                        batch_data = cursor.fetchall()
                        if not batch_data:
                            logger.info("Достигнут конец данных")
                            break

                        # Параллельная обработка пакета
                        chunk_size = len(batch_data) // num_processes + 1
                        chunks = [
                            batch_data[i : i + chunk_size]
                            for i in range(0, len(batch_data), chunk_size)
                        ]
                        chunk_data = [(chunk, ref_data, 0.6) for chunk in chunks]

                        logger.info(
                            f"Начало обработки пакета {processed_records + 1}-{processed_records + len(batch_data)} из {total_records} записей"
                        )

                        # Создаем пул процессов с тайм-аутом
                        with mp.Pool(processes=num_processes) as pool:
                            chunk_results = []
                            for result in pool.imap_unordered(
                                process_chunk, chunk_data
                            ):
                                chunk_results.extend(result)
                                logger.info(
                                    f"Получены результаты для части пакета, найдено {len(result)} совпадений"
                                )

                            all_results.extend(chunk_results)
                            logger.info(
                                f"Пакет обработан, всего найдено {len(chunk_results)} совпадений"
                            )

                        processed_records += len(batch_data)
                        logger.info(
                            f"Обработано {processed_records} записей из {total_records} ({(processed_records/total_records*100):.1f}%)"
                        )

                    except Exception as e:
                        logger.error(
                            f"Ошибка при обработке пакета: {str(e)}", exc_info=True
                        )
                        # Пропускаем проблемный пакет и продолжаем с следующего
                        processed_records += BATCH_SIZE
                        continue

                logger.info(
                    f"Обработка завершена. Всего найдено {len(all_results)} совпадений"
                )

                # Сохраняем результаты в сессии
                request.session["analysis_results"] = all_results

                # Сортируем результаты
                all_results.sort(
                    key=lambda x: float(x["similarity"].rstrip("%")), reverse=True
                )
                logger.info("Результаты отсортированы по убыванию процента схожести")

                return JsonResponse({"results": all_results})

            except Exception as e:
                logger.error(
                    f"Ошибка при работе с базой данных: {str(e)}", exc_info=True
                )
                return JsonResponse(
                    {"error": f"Ошибка при работе с базой данных: {str(e)}"}, status=500
                )

    except Exception as e:
        logger.error(f"Критическая ошибка в analyze_numbers: {str(e)}", exc_info=True)
        return JsonResponse({"error": f"Критическая ошибка: {str(e)}"}, status=500)


@staff_member_required
def replace_numbers(request):
    if request.method != "POST":
        return JsonResponse({"error": "Метод не поддерживается"}, status=405)

    try:
        data = json.loads(request.body)
        replacements = data.get("replacements", [])

        with connection.cursor() as cursor:
            for item in replacements:
                cursor.execute(
                    f"""
                    UPDATE {POTOK_TABLE}
                    SET gosnmr = %s
                    WHERE potok_id = %s
                """,
                    (item["suggested"], item["id"]),
                )

            connection.commit()

        return JsonResponse(
            {"success": True, "message": f"Обновлено {len(replacements)} записей"}
        )

    except Exception as e:
        logger.error(f"Ошибка при замене номеров: {str(e)}")
        return JsonResponse({"error": str(e)}, status=500)


def process_plate(plate):
    plate = str(plate).upper().strip()

    # Находим последнюю цифру и обрезаем все после нее
    last_digit_match = re.search(r"\d(?!.*\d)", plate)
    if last_digit_match:
        end_pos = last_digit_match.end()
        plate = plate[:end_pos]

    # Проверяем наличие английских букв
    eng_letters = set("ABCDEFGHIJKLMNOPQRSTUVWXYZ")
    if not any(c in eng_letters for c in plate):
        return plate

    en_to_ru = {
        "A": "А",
        "B": "В",
        "E": "Е",
        "K": "К",
        "M": "М",
        "H": "Н",
        "O": "О",
        "P": "Р",
        "C": "С",
        "T": "Т",
        "X": "Х",
        "Y": "У",
    }

    # Паттерны российских номеров
    patterns = [
        r"^[АВЕКМНОРСТУХ]\d{3}[АВЕКМНОРСТУХ]{2}\d{2,3}$",  # Стандартный
        r"^\d{4}[АВЕКМНОРСТУХ]{2}\d{2,3}$",  # Прицеп
        r"^[АВЕКМНОРСТУХ]\d{3}[АВЕКМНОРСТУХ]{2}$",  # Дипломатические
        r"^[АВЕКМНОРСТУХ]{2}\d{5,7}$",  # Спецтранспорт
    ]

    # Заменяем английские буквы на русские для проверки
    test_plate = plate
    for en, ru in en_to_ru.items():
        test_plate = test_plate.replace(en, ru)

    is_russian = any(re.match(pattern, test_plate) for pattern in patterns)

    if is_russian:
        for en, ru in en_to_ru.items():
            plate = plate.replace(en, ru)

    return plate


def get_schema_name():
    with connection.cursor() as cursor:
        cursor.execute("SELECT current_schema()")
        return cursor.fetchone()[0]


@staff_member_required
def upload_potok(request):
    if request.method == "POST":
        form = PotokUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                df = pd.read_excel(
                    request.FILES["file"],
                    engine="openpyxl",
                    dtype={
                        "Гос. номер": str,
                        "Камера": str,
                        "Приближение/удаление": str,
                    },
                )

                if len(pd.ExcelFile(request.FILES["file"]).sheet_names) > 1:
                    messages.error(
                        request, "Excel файл должен содержать только один лист"
                    )
                    return redirect("admin:index")

                required_headers = {
                    "A": "Дата фиксации",
                    "B": "Гос. номер",
                    "D": "Приближение/удаление",
                    "E": "Камера",
                }

                for col, header in required_headers.items():
                    if df.columns[ord(col) - ord("A")] != header:
                        messages.error(
                            request,
                            f"Неверный заголовок в столбце {col}. Ожидается: {header}",
                        )
                        return redirect("admin:index")

                # Оптимизированная обработка дат
                df["Дата фиксации"] = pd.to_datetime(
                    df["Дата фиксации"], format="mixed"
                )

                # Получаем даты из формы и убираем информацию о часовом поясе
                date_from = form.cleaned_data["date_from"]
                if date_from.tzinfo:
                    date_from = date_from.replace(tzinfo=None)

                date_to = form.cleaned_data["date_to"]
                if date_to.tzinfo:
                    date_to = date_to.replace(tzinfo=None)
                date_to = date_to.replace(hour=23, minute=59, second=59)

                # Проверяем диапазон дат
                dates_mask = (df["Дата фиксации"] >= date_from) & (
                    df["Дата фиксации"] <= date_to
                )
                if not dates_mask.all():
                    messages.error(
                        request, "В файле есть записи вне выбранного диапазона дат"
                    )
                    return redirect("admin:index")

                # Векторизованная обработка данных
                df["Гос. номер"] = df["Гос. номер"].apply(process_plate)
                df["Приближение/удаление"] = df["Приближение/удаление"].str.upper()
                df["Приближение/удаление"] = (
                    df["Приближение/удаление"]
                    .replace({"": None, "NAN": None})
                    .where(pd.notna(df["Приближение/удаление"]), None)
                )

                schema_name = get_schema_name()

                from psycopg2.extras import execute_values

                with connection.cursor() as cursor:
                    # Получаем максимальный potok_id одним запросом
                    cursor.execute(
                        f"SELECT COALESCE(MAX(potok_id), 0) FROM {POTOK_TABLE}"
                    )
                    max_potok_id = cursor.fetchone()[0]

                    # Устанавливаем таймзону
                    cursor.execute("SET timezone TO 'Asia/Yekaterinburg'")

                    # Подготавливаем данные для пакетной вставки
                    values = []
                    filename = request.FILES["file"].name

                    for idx, row in df.iterrows():
                        values.append(
                            (
                                max_potok_id + idx + 1,  # potok_id
                                row["Гос. номер"],  # gosnmr
                                row["Дата фиксации"],  # dt
                                row["Камера"],  # camera
                                row["Приближение/удаление"],  # direction
                                filename,  # filename
                                date_from,  # date_load_bgn
                                date_to,  # date_load_end
                            )
                        )

                    # Выполняем пакетную вставку
                    insert_query = f"""
                        INSERT INTO {POTOK_TABLE} (
                            potok_id, gosnmr, dt, camera, direction, filename,
                            date_load_bgn, date_load_end
                        ) VALUES %s
                    """

                    # Используем размер пакета 5000 записей
                    batch_size = 5000
                    for i in range(0, len(values), batch_size):
                        batch = values[i : i + batch_size]
                        execute_values(cursor, insert_query, batch)

                    connection.commit()

                messages.success(request, "Файл успешно обработан")
                return redirect("admin:index")

            except Exception as e:
                messages.error(request, f"Ошибка при обработке файла: {str(e)}")
                return redirect("admin:index")
    else:
        form = PotokUploadForm()

    return render(request, "admin/upload_potok.html", {"form": form})


def get_next_id(cursor, table, id_field):
    """Получает следующий ID для указанной таблицы"""
    cursor.execute(f"SELECT MAX({id_field}) FROM {table}")
    max_id = cursor.fetchone()[0]
    return (max_id or 0) + 1


def get_next_company_id(cursor):
    """Получает следующий company_id более надежным способом"""
    cursor.execute(
        f"""
        SELECT company_id 
        FROM {COMPANY_TABLE} 
        ORDER BY company_id DESC 
        LIMIT 1
    """
    )
    result = cursor.fetchone()
    return (result[0] if result else 0) + 1


def process_dates(date_value):
    """Обрабатывает значения дат, возвращая None для некорректных значений"""
    if pd.isna(date_value) or date_value is None or isinstance(date_value, str):
        return None
    return date_value.date() if isinstance(date_value, pd.Timestamp) else None


def process_mass(mass_str):
    """Извлекает числовое значение массы из строки"""
    if pd.isna(mass_str):
        return None
    # Если передано число, просто возвращаем его как int
    if isinstance(mass_str, (int, float)):
        return int(mass_str)
    # Извлекаем только цифры из строки
    mass_digits = re.findall(r"\d+", str(mass_str))
    return int(mass_digits[0]) if mass_digits else None


@staff_member_required
def upload_propusk(request):
    if request.method == "POST":
        form = PropuskUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                df = pd.read_excel(request.FILES["file"])
                propusk_type = form.cleaned_data["propusk_type"]

                # Получаем схему базы данных
                schema_name = get_schema_name()

                with connection.cursor() as cursor:
                    # Отключаем индексы перед массовой вставкой
                    cursor.execute(
                        f"""
                        ALTER TABLE {PROPUSK_TABLE} SET UNLOGGED;
                        DROP INDEX IF EXISTS idx_{PROPUSK_TABLE}_gn;
                        DROP INDEX IF EXISTS idx_{PROPUSK_TABLE}_company_id;
                        DROP INDEX IF EXISTS idx_{PROPUSK_TABLE}_dateactual;
                    """
                    )

                    # Устанавливаем таймзону
                    cursor.execute("SET timezone TO 'Asia/Yekaterinburg'")

                    # Получаем следующий propusk_id одним запросом
                    cursor.execute(
                        f"SELECT COALESCE(MAX(propusk_id), 0) FROM {PROPUSK_TABLE}"
                    )
                    next_propusk_id = cursor.fetchone()[0] + 1

                    # Получаем следующий company_id одним запросом
                    next_company_id = get_next_company_id(cursor)

                    # Получаем следующий companynr одним запросом
                    cursor.execute(
                        f"SELECT COALESCE(MAX(companynr), 0) FROM {COMPANY_TABLE}"
                    )
                    next_companynr = cursor.fetchone()[0] + 1

                    # Кэшируем существующие компании
                    cursor.execute(
                        f"""
                        SELECT c.companynr, c.company_id, c.company
                        FROM {COMPANY_TABLE} c
                        WHERE c.del = 0 AND c.dateactual IS NULL
                    """
                    )
                    existing_companies = {
                        str(row[0]): {"id": row[1], "name": row[2]}
                        for row in cursor.fetchall()
                    }

                    # Подготавливаем данные для пакетной вставки
                    propusk_values = []
                    company_values = []
                    company_updates = []
                    BATCH_SIZE = (
                        500  # Уменьшаем размер пакета для более стабильной работы
                    )

                    if propusk_type == "kronos":
                        # Проверяем заголовки для базы Кронос
                        required_headers = {
                            "C": "№ ПРОПУСКА",
                            "K": "автомобиль",
                            "L": "Гос.номер",
                            "N": "Разрешенная макс. масса (КГ)ТС",
                            "P": "Срок действия (с..)",
                            "Q": "Срок действия (по..)",
                            "T": "Продлён до",
                            "V": "Комментарий",
                            "W": "Договорные отношения",
                        }

                        for col, header in required_headers.items():
                            if df.columns[ord(col) - ord("A")] != header:
                                messages.error(
                                    request,
                                    f"Неверный заголовок в столбце {col}. Ожидается: {header}",
                                )
                                return redirect("admin:index")

                        # Преобразуем даты
                        date_columns = [
                            "Срок действия (с..)",
                            "Срок действия (по..)",
                            "Продлён до",
                        ]
                        for col in date_columns:
                            df[col] = pd.to_datetime(
                                df[col], format="%d.%m.%Y", errors="coerce"
                            )
                            df[col] = df[col].where(pd.notna(df[col]), None)

                        # Обрабатываем данные пакетами
                        for idx, row in df.iterrows():
                            company_str = str(row["От кого письмо"])

                            # Получаем информацию о компании
                            (
                                action_type,
                                existing_id,
                                companynr,
                                company_name,
                                old_id,
                            ) = get_company_info(cursor, company_str)

                            if action_type == "use_existing":
                                company_id = existing_id
                            elif action_type == "reactivate":
                                # Деактивируем активные записи
                                cursor.execute(
                                    f"""
                                    UPDATE {COMPANY_TABLE}
                                    SET dateactual = CURRENT_DATE
                                    WHERE companynr = %s AND dateactual IS NULL AND del = 0
                                    """,
                                    [companynr],
                                )

                                # Получаем pid от последней записи
                                cursor.execute(
                                    f"""
                                    SELECT company_id 
                                    FROM {COMPANY_TABLE}
                                    WHERE companynr = %s AND del = 0
                                    ORDER BY company_id DESC
                                    LIMIT 1
                                    """,
                                    [companynr],
                                )
                                last_record = cursor.fetchone()
                                old_company_id = last_record[0] if last_record else None

                                # Реактивируем запись с pid от последней записи
                                cursor.execute(
                                    f"""
                                    UPDATE {COMPANY_TABLE}
                                    SET dateactual = NULL, pid = %s
                                    WHERE company_id = %s
                                    """,
                                    [old_company_id, existing_id],
                                )
                                company_id = existing_id
                            else:  # create_new
                                # Деактивируем активные записи
                                cursor.execute(
                                    f"""
                                    UPDATE {COMPANY_TABLE}
                                    SET dateactual = CURRENT_DATE
                                    WHERE companynr = %s AND dateactual IS NULL AND del = 0
                                    """,
                                    [companynr],
                                )

                                # Получаем pid от последней записи
                                cursor.execute(
                                    f"""
                                    SELECT company_id 
                                    FROM {COMPANY_TABLE}
                                    WHERE companynr = %s AND del = 0
                                    ORDER BY company_id DESC
                                    LIMIT 1
                                    """,
                                    [companynr],
                                )
                                last_record = cursor.fetchone()
                                old_company_id = last_record[0] if last_record else None

                                # Создаем новую запись с pid от последней записи
                                cursor.execute(
                                    f"""
                                    INSERT INTO {COMPANY_TABLE} (company_id, company, del, pid, companynr)
                                    VALUES (%s, %s, 0, %s, %s)
                                    """,
                                    [
                                        next_company_id,
                                        company_name,
                                        old_company_id,
                                        companynr,
                                    ],
                                )
                                company_id = next_company_id
                                next_company_id += 1

                            # Добавляем запись пропуска
                            propusk_values.append(
                                (
                                    next_propusk_id + idx,
                                    process_plate(
                                        str(row["Гос.номер"]).replace(" ", "")
                                    ),
                                    company_id,
                                    process_dates(row["Срок действия (с..)"]),
                                    process_dates(row["Срок действия (по..)"]),
                                    row["№ ПРОПУСКА"],
                                    row["Договорные отношения"],
                                    2,  # tct_id для Кронос
                                    process_mass(row["Разрешенная макс. масса (КГ)ТС"]),
                                    row["автомобиль"],
                                    process_dates(row["Продлён до"]),
                                    (
                                        row["Комментарий"]
                                        if pd.notna(row["Комментарий"])
                                        else None
                                    ),
                                )
                            )

                            # Выполняем пакетную вставку при достижении размера пакета
                            if len(propusk_values) >= BATCH_SIZE:
                                execute_batch_insert(
                                    cursor,
                                    company_updates,
                                    company_values,
                                    propusk_values,
                                )
                                company_updates = []
                                company_values = []
                                propusk_values = []

                    elif propusk_type == "razoviy":
                        # Проверяем заголовки для разовых пропусков
                        required_headers = {
                            "A": "Контрагент",
                            "B": "Госномер",
                            "C": "Дата начала проезда",
                            "D": "Дата окончания проезда",
                            "E": "Марка",
                            "F": "Масса авто",
                            "G": "Номер пропуска",
                            "H": "Зона покрытия",
                            "I": "Договорные отношения",
                            "J": "Продлен до",
                            "K": "Комментарий",
                        }

                        for col, header in required_headers.items():
                            if df.columns[ord(col) - ord("A")] != header:
                                messages.error(
                                    request,
                                    f"Неверный заголовок в столбце {col}. Ожидается: {header}",
                                )
                                return redirect("admin:index")

                        # Пропускаем первые две строки
                        df = df.iloc[1:]

                        # Проверяем и пропускаем пустые строки
                        df = df.dropna(subset=["Госномер", "Контрагент"], how="all")
                        df = df[df["Госномер"].astype(str).str.strip() != ""]
                        df = df[df["Контрагент"].astype(str).str.strip() != ""]

                        # Преобразуем даты с явным указанием формата и обработкой ошибок
                        date_columns = [
                            "Дата начала проезда",
                            "Дата окончания проезда",
                            "Продлен до",
                        ]

                        for col in date_columns:
                            try:
                                # Сначала очищаем данные
                                df[col] = df[col].astype(str).str.strip()
                                df[col] = df[col].replace(
                                    ["nan", "NaN", "", "None"], pd.NaT
                                )

                                # Пробуем разные форматы дат
                                def parse_date(x):
                                    if pd.isna(x) or x is pd.NaT:
                                        return pd.NaT
                                    try:
                                        # Пробуем стандартный формат
                                        return pd.to_datetime(x, format="%d.%m.%Y")
                                    except:
                                        try:
                                            # Пробуем автоматическое определение формата
                                            return pd.to_datetime(x)
                                        except:
                                            logger.warning(
                                                f"Не удалось преобразовать дату: {x}"
                                            )
                                            return pd.NaT

                                df[col] = df[col].apply(parse_date)
                                logger.info(
                                    f"Обработан столбец {col}. Количество валидных дат: {df[col].notna().sum()}"
                                )
                            except Exception as e:
                                logger.error(
                                    f"Ошибка при обработке столбца {col}: {str(e)}"
                                )
                                df[col] = pd.NaT

                        # Обрабатываем данные пакетами
                        for idx, row in df.iterrows():
                            try:
                                company_str = (
                                    str(row["Контрагент"])
                                    if pd.notna(row["Контрагент"])
                                    else ""
                                )
                                # Получаем информацию о компании
                                (
                                    action_type,
                                    existing_id,
                                    companynr,
                                    company_name,
                                    old_id,
                                ) = get_company_info(cursor, company_str)

                                if action_type == "use_existing":
                                    company_id = existing_id
                                elif action_type == "reactivate":
                                    # Деактивируем активные записи
                                    cursor.execute(
                                        f"""
                                        UPDATE {COMPANY_TABLE}
                                        SET dateactual = CURRENT_DATE
                                        WHERE companynr = %s AND dateactual IS NULL AND del = 0
                                        """,
                                        [companynr],
                                    )

                                    # Получаем pid от последней записи
                                    cursor.execute(
                                        f"""
                                        SELECT company_id 
                                        FROM {COMPANY_TABLE}
                                        WHERE companynr = %s AND del = 0
                                        ORDER BY company_id DESC
                                        LIMIT 1
                                        """,
                                        [companynr],
                                    )
                                    last_record = cursor.fetchone()
                                    old_company_id = (
                                        last_record[0] if last_record else None
                                    )

                                    # Реактивируем запись с pid от последней записи
                                    cursor.execute(
                                        f"""
                                        UPDATE {COMPANY_TABLE}
                                        SET dateactual = NULL, pid = %s
                                        WHERE company_id = %s
                                        """,
                                        [old_company_id, existing_id],
                                    )
                                    company_id = existing_id
                                else:  # create_new
                                    # Деактивируем активные записи
                                    cursor.execute(
                                        f"""
                                        UPDATE {COMPANY_TABLE}
                                        SET dateactual = CURRENT_DATE
                                        WHERE companynr = %s AND dateactual IS NULL AND del = 0
                                        """,
                                        [companynr],
                                    )

                                    # Получаем pid от последней записи
                                    cursor.execute(
                                        f"""
                                        SELECT company_id 
                                        FROM {COMPANY_TABLE}
                                        WHERE companynr = %s AND del = 0
                                        ORDER BY company_id DESC
                                        LIMIT 1
                                        """,
                                        [companynr],
                                    )
                                    last_record = cursor.fetchone()
                                    old_company_id = (
                                        last_record[0] if last_record else None
                                    )

                                    # Создаем новую запись с pid от последней записи
                                    cursor.execute(
                                        f"""
                                        INSERT INTO {COMPANY_TABLE} (company_id, company, del, pid, companynr)
                                        VALUES (%s, %s, 0, %s, %s)
                                        """,
                                        [
                                            next_company_id,
                                            company_name,
                                            old_company_id,
                                            companynr,
                                        ],
                                    )
                                    company_id = next_company_id
                                    next_company_id += 1

                                # Добавляем запись пропуска
                                propusk_values.append(
                                    (
                                        next_propusk_id + idx,
                                        process_plate(
                                            str(row["Госномер"]).replace(" ", "")
                                            if pd.notna(row["Госномер"])
                                            else ""
                                        ),
                                        company_id,
                                        process_dates(row["Дата начала проезда"]),
                                        process_dates(row["Дата окончания проезда"]),
                                        (
                                            str(row["Номер пропуска"])
                                            if pd.notna(row["Номер пропуска"])
                                            else None
                                        ),
                                        (
                                            clean_company_name(
                                                str(row["Договорные отношения"])
                                            )
                                            if pd.notna(row["Договорные отношения"])
                                            else None
                                        ),
                                        3,  # tct_id для разовых пропусков
                                        process_mass(row["Масса авто"]),
                                        (
                                            str(row["Марка"])
                                            if pd.notna(row["Марка"])
                                            else None
                                        ),
                                        process_dates(row["Продлен до"]),
                                        (
                                            str(row["Комментарий"])
                                            if pd.notna(row["Комментарий"])
                                            else None
                                        ),
                                    )
                                )
                            except Exception as e:
                                logger.error(
                                    f"Ошибка при обработке строки {idx}: {str(e)}"
                                )
                                continue

                            # Выполняем пакетную вставку при достижении размера пакета
                            if len(propusk_values) >= BATCH_SIZE:
                                execute_batch_insert(
                                    cursor,
                                    company_updates,
                                    company_values,
                                    propusk_values,
                                )
                                company_updates = []
                                company_values = []
                                propusk_values = []

                    elif propusk_type == "gdya":
                        # Проверяем заголовки для списка ГДЯ
                        required_headers = {
                            "A": "№ пропуска",
                            "C": "Транспортное средство",
                            "D": "Гос. номер",
                            "E": "Получатель пропуска",
                            "G": "Срок действия",
                            "I": "Дата оформления",
                        }

                        for col, header in required_headers.items():
                            if df.columns[ord(col) - ord("A")] != header:
                                messages.error(
                                    request,
                                    f"Неверный заголовок в столбце {col}. Ожидается: {header}",
                                )
                                return redirect("admin:index")

                        # Преобразуем даты
                        date_columns = ["Срок действия", "Дата оформления"]
                        for col in date_columns:
                            df[col] = pd.to_datetime(
                                df[col], format="%d.%m.%Y", errors="coerce"
                            )

                        # Получаем company_id для ГДЯ (companynr = 200) из кэша
                        gdya_company_id = None
                        if "200" in existing_companies:
                            gdya_company_id = existing_companies["200"]["id"]
                        else:
                            # Если нет в кэше, создаем новую запись
                            company_values.append(
                                (next_company_id, "ГДЯ", 0, None, 200)
                            )
                            gdya_company_id = next_company_id
                            next_company_id += 1
                            existing_companies["200"] = {
                                "id": gdya_company_id,
                                "name": "ГДЯ",
                            }

                        if not gdya_company_id:
                            messages.error(
                                request,
                                "Не найдена запись компании ГДЯ (companynr = 200)",
                            )
                            return redirect("admin:index")

                        # Обрабатываем данные пакетами
                        for idx, row in df.iterrows():
                            # Добавляем запись пропуска
                            propusk_values.append(
                                (
                                    next_propusk_id + idx,
                                    process_plate(
                                        str(row["Гос. номер"]).replace(" ", "")
                                    ),
                                    gdya_company_id,
                                    process_dates(row["Дата оформления"]),
                                    process_dates(row["Срок действия"]),
                                    row["№ пропуска"],
                                    row["Получатель пропуска"],
                                    5,  # tct_id для списка ГДЯ
                                    0,  # mass = 0
                                    row["Транспортное средство"],
                                    None,  # prodlen
                                    None,  # coment
                                )
                            )

                            # Выполняем пакетную вставку при достижении размера пакета
                            if len(propusk_values) >= BATCH_SIZE:
                                execute_batch_insert(
                                    cursor,
                                    company_updates,
                                    company_values,
                                    propusk_values,
                                )
                                company_updates = []
                                company_values = []
                                propusk_values = []

                    # Выполняем оставшиеся вставки
                    if company_updates or company_values or propusk_values:
                        execute_batch_insert(
                            cursor, company_updates, company_values, propusk_values
                        )

                    # Восстанавливаем индексы
                    cursor.execute(
                        f"""
                        ALTER TABLE {PROPUSK_TABLE} SET LOGGED;
                        CREATE INDEX IF NOT EXISTS idx_{PROPUSK_TABLE}_gn ON {PROPUSK_TABLE}(gn);
                        CREATE INDEX IF NOT EXISTS idx_{PROPUSK_TABLE}_company_id ON {PROPUSK_TABLE}(company_id);
                        CREATE INDEX IF NOT EXISTS idx_{PROPUSK_TABLE}_dateactual ON {PROPUSK_TABLE}(dateactual);
                        ANALYZE {PROPUSK_TABLE};
                    """
                    )

                    connection.commit()
                    messages.success(request, "Файл успешно обработан")
                    return redirect("admin:index")

            except Exception as e:
                messages.error(request, f"Ошибка при обработке файла: {str(e)}")
                return redirect("admin:index")
    else:
        form = PropuskUploadForm()

    return render(request, "admin/upload_propusk.html", {"form": form})


def execute_batch_insert(cursor, company_updates, company_values, propusk_values):
    """Выполняет пакетную вставку данных"""
    if company_updates:
        execute_values(
            cursor,
            f"""
            UPDATE {COMPANY_TABLE}
            SET dateactual = data.dateactual
            FROM (VALUES %s) AS data(dateactual, company_id)
            WHERE {COMPANY_TABLE}.company_id = data.company_id
        """,
            company_updates,
        )

    if company_values:
        execute_values(
            cursor,
            f"""
            INSERT INTO {COMPANY_TABLE} (company_id, company, del, pid, companynr)
            VALUES %s
        """,
            company_values,
        )

    if propusk_values:
        execute_values(
            cursor,
            f"""
            INSERT INTO {PROPUSK_TABLE} (
                propusk_id, gn, company_id, dateb, datee, num,
                contractrelationship, tct_id, mass, marka,
                prodlen, coment
            ) VALUES %s
        """,
            propusk_values,
        )


def get_company_info(cursor, company_str):
    """
    Анализирует информацию о компании и возвращает кортеж:
    (
        action_type: str,  # Тип действия: 'use_existing', 'reactivate', 'create_new'
        company_id: int | None,  # ID существующей компании или None
        companynr: int,  # Номер компании
        company_name: str,  # Очищенное название компании
        old_company_id: int | None  # ID старой записи для обновления (если есть)
    )
    """
    company_str = company_str.strip()
    company_name = None
    companynr = 0

    # Извлекаем companynr и название из строки
    match = re.match(r"^(\d+)\.(.+)$", company_str)
    if match:
        companynr = int(match.group(1))
        company_name = " ".join(match.group(2).split())
    else:
        company_name = " ".join(company_str.split())
        companynr = 0

    if companynr == 0:
        # Для companynr = 0 ищем только по имени
        cursor.execute(
            f"""
            SELECT company_id, company, dateactual, companynr
            FROM {COMPANY_TABLE}
            WHERE company = %s AND del = 0
            ORDER BY dateactual NULLS FIRST, company_id DESC
            """,
            [company_name],
        )
        existing = cursor.fetchone()

        if existing:
            company_id, existing_name, dateactual, existing_companynr = existing
            existing_name = " ".join(existing_name.split())

            # Сценарий 1: Найдена активная запись
            if dateactual is None and existing_name == company_name:
                return (
                    "use_existing",
                    company_id,
                    existing_companynr,
                    company_name,
                    None,
                )

            # Сценарий 2: Найдена неактивная запись
            if dateactual is not None and existing_name == company_name:
                return (
                    "reactivate",
                    company_id,
                    existing_companynr,
                    company_name,
                    None,
                )

        # Сценарий 3: Запись не найдена
        cursor.execute(
            f"""
            SELECT companynr + 1
            FROM {COMPANY_TABLE}
            WHERE companynr NOT IN (0, 1)
            ORDER BY companynr DESC
            LIMIT 1
            """
        )
        new_companynr = cursor.fetchone()[0]
        return ("create_new", None, new_companynr, company_name, None)

    elif companynr == 1:
        # Для companynr = 1 всегда берем значения из активной системной записи
        cursor.execute(
            f"""
            SELECT company_id, company, companynr
            FROM {COMPANY_TABLE}
            WHERE companynr = 1 AND dateactual IS NULL AND del = 0
            """
        )
        system_record = cursor.fetchone()
        if system_record:
            return (
                "use_existing",
                system_record[0],
                system_record[2],
                system_record[1],
                None,
            )
        else:
            raise ValueError(
                "Системная запись с companynr = 1 не найдена или неактивна"
            )

    else:
        # Для companynr ≠ 0 ищем по companynr
        cursor.execute(
            f"""
            SELECT company_id, company, dateactual, companynr
            FROM {COMPANY_TABLE}
            WHERE companynr = %s AND del = 0
            ORDER BY dateactual NULLS FIRST, company_id DESC
            """,
            [companynr],
        )
        existing = cursor.fetchone()

        if existing:
            company_id, existing_name, dateactual, existing_companynr = existing
            existing_name = " ".join(existing_name.split())

            # Сценарий 1: Найдена активная запись с тем же именем
            if dateactual is None and existing_name == company_name:
                return (
                    "use_existing",
                    company_id,
                    existing_companynr,
                    company_name,
                    None,
                )

            # Сценарий 2: Найдена неактивная запись с тем же именем
            if dateactual is not None and existing_name == company_name:
                return (
                    "reactivate",
                    company_id,
                    existing_companynr,
                    company_name,
                    None,
                )

        # Сценарий 3: Найдена запись с другим именем или запись не найдена
        return ("create_new", None, companynr, company_name, None)


def get_vehicle_category(mass):
    """Определяет категорию ТС на основе массы"""
    if mass is None:
        return None
    if mass <= 3500:
        return 1
    elif mass <= 10000:
        return 2
    elif mass <= 25000:
        return 3
    else:
        return 4


@staff_member_required
def generate_report(request):
    if request.method == "POST":
        try:
            # Получаем дату и тарифы из формы
            report_date = datetime.strptime(request.POST.get("report_date"), "%Y-%m")
            month_name = report_date.strftime("%B_%Y")

            month_translations = {
                "January": "Январь",
                "February": "Февраль",
                "March": "Март",
                "April": "Апрель",
                "May": "Май",
                "June": "Июнь",
                "July": "Июль",
                "August": "Август",
                "September": "Сентябрь",
                "October": "Октябрь",
                "November": "Ноябрь",
                "December": "Декабрь",
            }

            # Переводим название месяца на русский
            month_eng = report_date.strftime("%B")
            month_rus = month_translations[month_eng]
            sheet_name = f"{month_rus}_{report_date.year}"

            # Получаем тарифы
            tariffs = {
                1: float(request.POST.get("tariff_1", 0)),
                2: float(request.POST.get("tariff_2", 0)),
                3: float(request.POST.get("tariff_3", 0)),
                4: float(request.POST.get("tariff_4", 0)),
            }

            # Создаем два Excel файла
            wb = Workbook()
            wb.remove(wb.active)  # Удаляем стандартный лист
            ws = wb.create_sheet(title=sheet_name)

            wb_simple = Workbook()
            wb_simple.remove(wb_simple.active)  # Удаляем стандартный лист
            ws_simple = wb_simple.create_sheet(title=sheet_name)

            # Создаем стили
            header_font = Font(bold=True)
            center_alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            red_fill = PatternFill(
                start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
            )
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Устанавливаем заголовки для полного отчета
            headers = [
                "Гос. номер",
                "№ фиксации",
                "Камера",
                "Дата фиксации",
                "Перевозчик",
                "Выбранный перевозчик",
                "Категория",
                "Тариф",
                "Сумма",
            ]

            # Устанавливаем заголовки для упрощенного отчета
            headers_simple = [
                "Гос. номер",
                "Кол-во фиксаций",
                "Перевозчик",
                "Выбранный перевозчик",
                "Сумма",
            ]

            # Заполняем заголовки и устанавливаем ширину столбцов
            for idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=idx)
                cell.value = header
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border

            for idx, header in enumerate(headers_simple, 1):
                cell = ws_simple.cell(row=1, column=idx)
                cell.value = header
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border

            # Устанавливаем ширину столбцов
            column_widths = {
                1: 15,
                2: 12,
                3: 20,
                4: 20,
                5: 30,
                6: 30,
                7: 15,
                8: 15,
                9: 15,
            }
            column_widths_simple = {1: 15, 2: 15, 3: 30, 4: 30, 5: 15}

            for col, width in column_widths.items():
                ws.column_dimensions[get_column_letter(col)].width = width

            for col, width in column_widths_simple.items():
                ws_simple.column_dimensions[get_column_letter(col)].width = width

            with connection.cursor() as cursor:
                # Получаем все действующие пропуска одним запросом
                cursor.execute(
                    f"""
                    WITH valid_numbers AS (
                        SELECT DISTINCT gn, mass, company_id, tct_id
                        FROM {PROPUSK_TABLE}
                        WHERE gn IS NOT NULL AND dateactual IS NULL
                    ),
                    company_info AS (
                        SELECT c.company_id, c.company, c.companynr
                        FROM {COMPANY_TABLE} c
                        WHERE c.del = 0 AND c.dateactual IS NULL
                    )
                    SELECT
                        vn.gn,
                        vn.mass,
                        vn.company_id,
                        vn.tct_id,
                        ci.company
                    FROM valid_numbers vn
                    LEFT JOIN company_info ci ON vn.company_id = ci.company_id
                    """
                )
                propusk_data = {
                    row[0].upper(): {
                        "mass": row[1],
                        "company_id": row[2],
                        "tct_id": row[3],
                        "company": row[4],
                    }
                    for row in cursor.fetchall()
                }

                # Получаем данные потока за выбранный месяц с группировкой
                cursor.execute(
                    f"""
                    WITH filtered_data AS (
                        SELECT
                            gosnmr,
                            camera,
                            direction,
                            dt,
                            LAG(dt) OVER (PARTITION BY gosnmr ORDER BY dt) as prev_dt
                        FROM {POTOK_TABLE}
                        WHERE
                            EXTRACT(YEAR FROM dt) = %s
                            AND EXTRACT(MONTH FROM dt) = %s
                            AND del IS NULL
                    ),
                    valid_records AS (
                        SELECT *
                        FROM filtered_data
                        WHERE prev_dt IS NULL OR dt - prev_dt >= interval '24 hours'
                    )
                    SELECT
                        gosnmr,
                        COUNT(*) as fixation_count,
                        array_agg(camera) as cameras,
                        array_agg(direction) as directions,
                        array_agg(dt) as dates
                    FROM valid_records
                    GROUP BY gosnmr
                    ORDER BY gosnmr
                    """,
                    [report_date.year, report_date.month],
                )

                current_row = 2
                current_row_simple = 2

                # Обрабатываем результаты
                for row in cursor.fetchall():
                    gosnmr, fixation_count, cameras, directions, dates = row
                    upper_gosnmr = gosnmr.upper()

                    # Получаем информацию о пропуске
                    propusk_info = propusk_data.get(upper_gosnmr, {})
                    mass = propusk_info.get("mass")
                    company = propusk_info.get("company", "")

                    # Определяем категорию и тариф
                    category = get_vehicle_category(mass)
                    tariff = tariffs.get(category, 0) if category else 0
                    total_sum = tariff * fixation_count if tariff else 0

                    # Записываем в полный отчет
                    for i in range(fixation_count):
                        # Гос. номер
                        cell = ws.cell(row=current_row + i, column=1)
                        cell.value = gosnmr
                        if i == 0:  # Только для первой строки номера
                            if upper_gosnmr not in propusk_data:
                                cell.fill = red_fill
                        cell.alignment = center_alignment
                        cell.border = thin_border

                        # № фиксации
                        ws.cell(
                            row=current_row + i, column=2, value=i + 1
                        ).alignment = center_alignment

                        # Камера
                        camera_info = (
                            f"{cameras[i]} ({directions[i]})"
                            if directions[i]
                            else cameras[i]
                        )
                        ws.cell(
                            row=current_row + i, column=3, value=camera_info
                        ).alignment = center_alignment

                        # Дата фиксации
                        ws.cell(
                            row=current_row + i,
                            column=4,
                            value=dates[i].strftime("%d.%m.%Y %H:%M:%S"),
                        ).alignment = center_alignment

                        # Остальные поля только для первой строки
                        if i == 0:
                            # Перевозчик
                            ws.cell(
                                row=current_row, column=5, value=company
                            ).alignment = center_alignment
                            # Выбранный перевозчик
                            ws.cell(
                                row=current_row, column=6, value=company
                            ).alignment = center_alignment
                            # Категория
                            ws.cell(
                                row=current_row, column=7, value=category
                            ).alignment = center_alignment
                            # Тариф
                            ws.cell(
                                row=current_row, column=8, value=f"{tariff:.2f}"
                            ).alignment = center_alignment
                            # Сумма
                            ws.cell(
                                row=current_row, column=9, value=f"{total_sum:.2f}"
                            ).alignment = center_alignment

                    # Объединяем ячейки для одинаковых значений
                    if fixation_count > 1:
                        for col in [1, 5, 6, 7, 8, 9]:
                            ws.merge_cells(
                                start_row=current_row,
                                start_column=col,
                                end_row=current_row + fixation_count - 1,
                                end_column=col,
                            )

                    # Записываем в упрощенный отчет
                    ws_simple.cell(
                        row=current_row_simple, column=1, value=gosnmr
                    ).alignment = center_alignment
                    if upper_gosnmr not in propusk_data:
                        ws_simple.cell(row=current_row_simple, column=1).fill = red_fill
                    ws_simple.cell(
                        row=current_row_simple, column=2, value=fixation_count
                    ).alignment = center_alignment
                    ws_simple.cell(
                        row=current_row_simple, column=3, value=company
                    ).alignment = center_alignment
                    ws_simple.cell(
                        row=current_row_simple, column=4, value=company
                    ).alignment = center_alignment
                    ws_simple.cell(
                        row=current_row_simple, column=5, value=f"{total_sum:.2f}"
                    ).alignment = center_alignment

                    # Добавляем границы для упрощенного отчета
                    for col in range(1, 6):
                        ws_simple.cell(row=current_row_simple, column=col).border = (
                            thin_border
                        )

                    current_row += fixation_count
                    current_row_simple += 1

            # Сохраняем файлы
            filename = f'Анализ_потока_{report_date.strftime("%m_%Y")}.xlsx'
            filename_simple = (
                f'Анализ_потока_краткий_{report_date.strftime("%m_%Y")}.xlsx'
            )
            filepath = os.path.join("media", filename)
            filepath_simple = os.path.join("media", filename_simple)

            os.makedirs("media", exist_ok=True)
            wb.save(filepath)
            wb_simple.save(filepath_simple)

            # Создаем ZIP-архив
            zip_filename = f'Анализ_потока_{report_date.strftime("%m_%Y")}.zip'
            zip_filepath = os.path.join("media", zip_filename)

            with zipfile.ZipFile(zip_filepath, "w") as zipf:
                zipf.write(filepath, os.path.basename(filepath))
                zipf.write(filepath_simple, os.path.basename(filepath_simple))

            # Отправляем ZIP-архив
            response = FileResponse(
                open(zip_filepath, "rb"),
                content_type="application/zip",
                as_attachment=True,
                filename=zip_filename,
            )

            # Добавляем callback для удаления файлов
            def cleanup():
                for f in [filepath, filepath_simple, zip_filepath]:
                    if os.path.exists(f):
                        os.remove(f)
                        logger.info(f"Временный файл {f} удален")

            response._resource_closers.append(cleanup)
            return response

        except Exception as e:
            # В случае ошибки удаляем файлы
            for f in (
                [filepath, filepath_simple, zip_filepath]
                if "filepath" in locals()
                else []
            ):
                if os.path.exists(f):
                    os.remove(f)
                    logger.info(f"Временный файл {f} удален после ошибки")
            messages.error(request, f"Ошибка при формировании отчета: {str(e)}")
            return redirect("admin:index")

    else:
        context = {
            "default_date": datetime.now().strftime("%Y-%m"),
            "default_tariffs": {
                "tariff_1": "1410,28",
                "tariff_2": "2820,57",
                "tariff_3": "5641,13",
                "tariff_4": "9025,81",
            },
        }
        return render(request, "admin/generate_report.html", context)


def clean_company_name(name):
    """Очищает название компании от номера в формате N."""
    if not name or not isinstance(name, str):
        return name
    match = re.match(r"^\d+\.(.*)", name.strip())
    return match.group(1).strip() if match else name.strip()
